import RPi.GPIO as GPIO
import time
import w1thermsensor
from openpyxl import Workbook
import datetime

#czujnik_temp
sensor = w1thermsensor.W1ThermSensor()

#skoroszyt
wb = Workbook()
sheet = wb.active
rowA =1
columnA =1
#data
data_rozp = datetime.datetime.now()
dzien = str(data_rozp.day)
miesiac = str(data_rozp.month)
rok = str(data_rozp.year)
nazwa_pliku = "temp_" + dzien + '.'+ miesiac + '.'+ rok + 'r'+".xlsx"
try:
    while True:
        temperatura = sensor.get_temperature()
        sheet.cell(row = rowA, column = columnA).value = temperatura
        #godzina
        aktualna_godzina = datetime.datetime.now()
        godz = str(aktualna_godzina.hour)
        minu = str(aktualna_godzina.minute)
        sheet.cell(row = rowA, column = columnA+1).value = godz + ':' + minu 
        
        rowA = rowA+1
        time.sleep(300)

except KeyboardInterrupt:
    wb.save(nazwa_pliku)
    print("Koniec. Utworzono: ", nazwa_pliku)
    #GPIO.cleanup()